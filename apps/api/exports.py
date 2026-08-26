from __future__ import annotations

import json
from html import escape
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


YAKU_NAMES = {
    "riichi": "立直",
    "double_riichi": "两立直",
    "ippatsu": "一发",
    "menzen_tsumo": "门前清自摸和",
    "tanyao": "断幺九",
    "pinfu": "平和",
    "iipeikou": "一杯口",
    "seat_wind_east": "自风东",
    "seat_wind_south": "自风南",
    "seat_wind_west": "自风西",
    "seat_wind_north": "自风北",
    "round_wind_east": "场风东",
    "round_wind_south": "场风南",
    "round_wind_west": "场风西",
    "round_wind_north": "场风北",
    "haku": "白",
    "hatsu": "发",
    "chun": "中",
    "rinshan": "岭上开花",
    "chankan": "抢杠",
    "haitei": "海底摸月",
    "houtei": "河底捞鱼",
    "sanshoku_doujun": "三色同顺",
    "ikkitsuukan": "一气通贯",
    "chanta": "混全带幺九",
    "chiitoitsu": "七对子",
    "toitoi": "对对和",
    "sanankou": "三暗刻",
    "honroutou": "混老头",
    "sanshoku_doukou": "三色同刻",
    "sankantsu": "三杠子",
    "shousangen": "小三元",
    "honitsu": "混一色",
    "junchan": "纯全带幺九",
    "ryanpeikou": "二杯口",
    "chinitsu": "清一色",
    "kokushi": "国士无双",
    "suuankou": "四暗刻",
    "daisangen": "大三元",
    "shousuushii": "小四喜",
    "daisuushii": "大四喜",
    "tsuuiisou": "字一色",
    "chinroutou": "清老头",
    "ryuuiisou": "绿一色",
    "chuuren": "九莲宝灯",
    "suukantsu": "四杠子",
    "tenhou": "天和",
    "chiihou": "地和",
    "renhou": "人和",
    "nagashi_mangan": "流局满贯",
    "dora": "宝牌",
    "ura_dora": "里宝牌",
    "aka_dora": "赤宝牌",
    "kita": "拔北",
    "double_yakuman": "双倍役满",
}


def _candidate_label(candidate: dict[str, Any]) -> str:
    if candidate.get("first_kyushu"):
        return "九種九牌流局"
    if candidate.get("first_kan"):
        prefix = "暗杠 "
    elif candidate.get("first_riichi"):
        prefix = "立直打 "
    else:
        prefix = "打 "
    return prefix + str(candidate.get("discard", ""))


def _get(value: Any, path: str) -> Any:
    for key in path.split("."):
        if isinstance(value, dict):
            value = value.get(key)
        elif isinstance(value, list) and key.isdigit() and int(key) < len(value):
            value = value[int(key)]
        else:
            return None
    return value


def _number(candidate: dict[str, Any], path: str) -> Any:
    value = _get(candidate, path)
    return value.get("value") if isinstance(value, dict) and "value" in value else value


def _ci(candidate: dict[str, Any], path: str) -> str:
    value = _get(candidate, path)
    ci95 = value.get("ci95") if isinstance(value, dict) else None
    if not ci95:
        return "样本不足"
    return f"{ci95[0]:+.2f} 至 {ci95[1]:+.2f}"


def _rank_ci(candidate: dict[str, Any]) -> str:
    value = _get(candidate, "hanchan.expected_rank")
    ci95 = value.get("ci95") if isinstance(value, dict) else None
    if not ci95:
        return "样本不足"
    return f"{ci95[0]:.3f} 至 {ci95[1]:.3f}"


def _rate(candidate: dict[str, Any], path: str) -> str:
    value = _get(candidate, path)
    if not isinstance(value, dict) or value.get("rate") is None:
        return "不可用"
    return f"{value['rate']:.2%} ({value.get('count', 0)}/{value.get('total', 0)})"


def _pair(candidate: dict[str, Any], left: str, right: str) -> str:
    left_value = _get(candidate, left)
    right_value = _get(candidate, right)
    if not isinstance(left_value, dict) or not isinstance(right_value, dict):
        return "不可用"
    if left_value.get("rate") is None or right_value.get("rate") is None:
        return "不可用"
    return f"{left_value['rate']:.2%} / {right_value['rate']:.2%}"


def _flatten(prefix: str, value: Any, output: list[tuple[str, Any]]) -> None:
    if isinstance(value, dict):
        for key, child in value.items():
            _flatten(f"{prefix}.{key}" if prefix else str(key), child, output)
    elif isinstance(value, list):
        output.append((prefix, json.dumps(value, ensure_ascii=False)))
    elif value is not None:
        output.append((prefix, value))


def _autosize(sheet: Any, maximum: int = 60) -> None:
    for index, column in enumerate(sheet.columns, 1):
        width = max((len(str(cell.value)) for cell in column if cell.value is not None), default=8)
        sheet.column_dimensions[get_column_letter(index)].width = min(maximum, max(10, width + 2))


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _append(sheet: Any, values: list[Any]) -> None:
    sheet.append([_excel_safe(value) for value in values])


def _style_header(cells: Any) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor="0B5D50")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")


def _metric_rows(
    candidates: list[dict[str, Any]],
    *,
    point_available: bool = True,
) -> list[tuple[str, str, Callable[[dict[str, Any]], Any]]]:
    rows: list[tuple[str, str, Callable[[dict[str, Any]], Any]]] = [
        ("核心指标", "平均局收支（NAGA 口径）", lambda c: _number(c, "value.point") if point_available else "不可用"),
        ("核心指标", "局收支 95% CI", lambda c: _ci(c, "value.point") if point_available else "不可用"),
        ("核心指标", "予想半荘終了時順位", lambda c: _number(c, "hanchan.expected_rank")),
        ("核心指标", "順位 95% CI", _rank_ci),
    ]
    rows.extend(
        ("核心指标", f"予想 {rank} 位率", lambda c, index=rank - 1: _rate(c, f"hanchan.rank_rates.{index}"))
        for rank in range(1, 5)
    )
    rows.extend(
        ("核心指标", f"予想終了時鳳{dan}段段位pt収支", lambda c, dan=dan: _number(c, f"hanchan.dan_pt_ev.houou_{dan}"))
        for dan in (7, 8, 9, 10)
    )
    rows.extend(
        ("核心指标", f"段位pt 95% CI（鳳{dan}段）", lambda c, dan=dan: _ci(c, f"hanchan.dan_pt_ev.houou_{dan}"))
        for dan in (7, 8, 9, 10)
    )
    rows.extend([
        ("终局分布", "自家和牌", lambda c: _rate(c, "outcome.self_win")),
        ("终局分布", "自家放铳", lambda c: _rate(c, "outcome.self_deal_in")),
        ("终局分布", "流局", lambda c: _rate(c, "outcome.draw")),
        ("终局分布", "横移动", lambda c: _rate(c, "outcome.sideways")),
        ("终局分布", "他家自摸", lambda c: _rate(c, "outcome.other_tsumo")),
        ("和牌构成与打点", "自摸 / 荣和", lambda c: _pair(c, "win.tsumo_share", "win.ron_share")),
        ("和牌构成与打点", "立直和牌占比", lambda c: _rate(c, "win.riichi_share")),
        ("和牌构成与打点", "默听和牌占比", lambda c: _rate(c, "win.dama_share")),
        ("和牌构成与打点", "副露和牌占比", lambda c: _rate(c, "win.open_share")),
        ("和牌构成与打点", "平均和牌点", lambda c: _number(c, "win.average_point")),
        ("和牌构成与打点", "平均和牌素点", lambda c: _number(c, "win.average_raw_point")),
        ("和牌构成与打点", "平均翻数", lambda c: _number(c, "win.average_han")),
        ("和牌构成与打点", "平均符数", lambda c: _number(c, "win.average_fu")),
        ("立直、听牌与副露", "立直率", lambda c: _rate(c, "riichi.rate")),
        ("立直、听牌与副露", "先制 / 追立占比", lambda c: _pair(c, "riichi.first_rate", "riichi.chase_rate")),
        ("立直、听牌与副露", "立直后和牌率", lambda c: _rate(c, "riichi.win_after_rate")),
        ("立直、听牌与副露", "平均立直巡目", lambda c: _number(c, "riichi.average_turn")),
        ("立直、听牌与副露", "听牌率", lambda c: _rate(c, "tenpai.rate")),
        ("立直、听牌与副露", "平均首次听牌巡目", lambda c: _number(c, "tenpai.average_first_turn")),
        ("立直、听牌与副露", "副露率", lambda c: _rate(c, "call.rate")),
        ("立直、听牌与副露", "平均副露数", lambda c: _number(c, "call.average_count")),
        ("立直、听牌与副露", "副露后和牌率", lambda c: _rate(c, "call.win_after_rate")),
        ("防守", "平均放铳损失", lambda c: _number(c, "defense.average_deal_in_loss")),
        ("防守", "平均放铳巡目", lambda c: _number(c, "defense.average_deal_in_turn")),
    ])
    yaku_ids = list(dict.fromkeys(item.get("id") for candidate in candidates for item in candidate.get("yaku", [])))
    for yaku_id in yaku_ids:
        def yaku_value(candidate: dict[str, Any], identifier: str = str(yaku_id)) -> str:
            item = next((entry for entry in candidate.get("yaku", []) if entry.get("id") == identifier), None)
            if not item or not item.get("available") or item.get("rate") is None:
                return "不可用"
            suffix = f" · 总枚数 {item['total_tiles']}" if item.get("total_tiles") is not None else ""
            return f"{item.get('count', 0)} 局 · {item['rate']:.2%}{suffix}"

        rows.append(("役种频率", f"{YAKU_NAMES.get(str(yaku_id), str(yaku_id))} [{yaku_id}]", yaku_value))
    return rows


def build_xlsx(result: dict[str, Any]) -> bytes:
    candidates = list(result.get("candidates") or result.get("summaries") or [])
    point_available = (result.get("metrics_version") or 0) >= 2
    workbook = Workbook()
    metrics = workbook.active
    metrics.title = "指标总表"
    metrics.freeze_panes = "C2"
    _append(metrics, ["分组", "指标", *(_candidate_label(candidate) for candidate in candidates)])
    _style_header(metrics[1])
    last_section = None
    for section, label, getter in _metric_rows(candidates, point_available=point_available):
        _append(metrics, [section if section != last_section else "", label, *(getter(candidate) for candidate in candidates)])
        last_section = section
    metrics.auto_filter.ref = metrics.dimensions
    _autosize(metrics)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _tile_class(tile: str) -> str:
    base = tile.strip().lower()
    if base.endswith("r"):
        base = base[:-1]
    if not base:
        return "t-z"
    if base.endswith("m"):
        return "t-man"
    if base.endswith("p"):
        return "t-pin"
    if base.endswith("s"):
        return "t-sou"
    return "t-z"


def _tile_chips(tiles: list[str]) -> str:
    return "".join(
        f"<span class='tile {_tile_class(tile)}'>{escape(tile)}</span>"
        for tile in tiles
    )


def _preset_html(result: dict[str, Any]) -> str:
    inp = result.get("resolved_input") or {}
    ctx = result.get("resolved_context") or {}
    candidates = list(result.get("candidates") or result.get("summaries") or [])
    hand = list(inp.get("main_haipai") or [])
    first_tsumo = inp.get("first_tsumo")
    dora = inp.get("dora")
    scores = ctx.get("scores") or []
    rel = ctx.get("relative_scores") or {}
    parts: list[str] = []

    parts.append("<div class='preset-grid'>")
    parts.append("<div class='preset-card'><div class='preset-title'>手牌（13 张）</div><div class='tiles'>")
    parts.append(_tile_chips(hand))
    parts.append("</div></div>")
    if first_tsumo:
        parts.append("<div class='preset-card'><div class='preset-title'>第一摸牌</div><div class='tiles'>")
        parts.append(_tile_chips([first_tsumo]))
        parts.append("</div></div>")
    if dora:
        parts.append("<div class='preset-card'><div class='preset-title'>宝牌指示</div><div class='tiles'>")
        parts.append(_tile_chips([dora]))
        parts.append("</div></div>")
    cand_text = "、".join(_candidate_label(c) for c in candidates) if candidates else "不可用"
    parts.append(f"<div class='preset-card'><div class='preset-title'>候选动作</div><div class='preset-value'>{escape(cand_text)}</div></div>")

    ctx_items = []
    if ctx.get("round"):
        ctx_items.append(("局", f"{ctx['round']}　{ctx.get('honba', 0)} 本场　{ctx.get('kyotaku', 0)} 供托"))
    if scores:
        ctx_items.append(("点数", "　".join(f"{s:,}" for s in scores)))
    elif rel:
        ctx_items.append(("点数", "　".join(f"{v:,}" for v in rel.values())))
    ctx_items.append(("局数", f"{result.get('runs') or result.get('total_runs') or '?'} 局"))
    if result.get("seed") is not None:
        ctx_items.append(("种子", str(result["seed"])))
    ctx_html = "".join(
        f"<div class='ctx-row'><span class='ctx-key'>{escape(k)}</span><span class='ctx-val'>{escape(str(v))}</span></div>"
        for k, v in ctx_items
    )
    parts.append(f"<div class='preset-card wide'><div class='preset-title'>模拟预设</div><div class='ctx'>{ctx_html}</div></div>")
    parts.append("</div>")
    return "".join(parts)


def build_html(result: dict[str, Any], run_id: str) -> str:
    candidates = list(result.get("candidates") or result.get("summaries") or [])
    headers = "".join(f"<th>{escape(_candidate_label(candidate))}</th>" for candidate in candidates)
    rows: list[str] = []
    last_section = None
    for section, label, getter in _metric_rows(candidates, point_available=(result.get("metrics_version") or 0) >= 2):
        if section != last_section:
            rows.append(f"<tr class='section'><th colspan='{len(candidates) + 1}'>{escape(section)}</th></tr>")
            last_section = section
        cells = []
        for candidate in candidates:
            value = getter(candidate)
            cells.append(f"<td>{escape(str(value if value is not None else '不可用'))}</td>")
        values = "".join(cells)
        rows.append(f"<tr><th>{escape(label)}</th>{values}</tr>")

    preset = _preset_html(result)
    return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width">
<title>MortalSim {escape(run_id)}</title>
<style>
body{{font:14px system-ui;margin:32px auto;max-width:1440px;padding:0 20px;color:#1d282d;background:#f4f7f6}}
h1,h2{{color:#123e36}}p{{color:#647276}}
.preset-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:14px;margin:18px 0 26px}}
.preset-card{{background:#fff;border:1px solid #d9e1de;border-radius:10px;padding:14px 16px;box-shadow:0 1px 3px rgba(0,0,0,.05)}}
.preset-card.wide{{grid-column:1/-1}}
.preset-title{{font-weight:700;color:#123e36;margin-bottom:10px;font-size:13px}}
.tiles{{display:flex;flex-wrap:wrap;gap:6px}}
.tile{{display:inline-flex;align-items:center;justify-content:center;min-width:34px;height:42px;padding:0 6px;border-radius:6px;border:1px solid #cfd8d5;background:linear-gradient(#fff,#f2f5f4);font-weight:700;font-size:15px;box-shadow:0 1px 2px rgba(0,0,0,.08)}}
.t-man{{color:#b03030}}.t-pin{{color:#1f5fb0}}.t-sou{{color:#1e8a4a}}.t-z{{color:#333}}
.preset-value{{font-size:15px;font-weight:600;color:#123e36}}
.ctx{{display:grid;grid-template-columns:auto 1fr;gap:6px 18px}}
.ctx-key{{color:#647276;font-weight:600}}.ctx-val{{color:#1d282d}}
.table-wrap{{overflow:auto;background:#fff;border:1px solid #d9e1de}}
table{{border-collapse:collapse;width:100%}}th,td{{padding:9px 12px;border-bottom:1px solid #e4e9e7;text-align:right;white-space:nowrap}}
th:first-child,td:first-child{{text-align:left;position:sticky;left:0;background:#fff}}thead th{{background:#123e36;color:#fff;position:sticky;top:0}}
.section th{{position:static!important;background:#e7f0ed!important;color:#264942!important;text-align:left!important}}
code{{font-family:ui-monospace,monospace}}
</style></head><body>
<h1>MortalSim 分析结果</h1><p>Run <code>{escape(run_id)}</code></p>
{preset}
<h2>候选指标总表</h2><div class="table-wrap"><table><thead><tr><th>指标</th>{headers}</tr></thead>
<tbody>{''.join(rows)}</tbody></table></div>
</body></html>"""
