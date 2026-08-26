from __future__ import annotations

from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest
from pydantic import ValidationError

import apps.api.main as main_module
import apps.api.job_manager as job_manager_module
from apps.api.job_manager import Job, JobManager
from apps.api.main import app
from apps.api.models import DiscardCandidate, RunRequest
from apps.api.services import StatisticsService
from mortal_app.service import _parse_inputs, resolve_simulation_context
from mortal_app.model_registry import ModelRegistry


FORMAL_RUNTIME = {
    "engine_id": "aoti-cuda-sm89",
    "artifact_sha256": "runtime-test",
    "build_id": "build-test",
    "compute_capability": "8.9",
    "batch_size": 1000,
    "batch_capacity": 1024,
    "precision_profile": "amp-static-advantage",
}


def test_empty_model_registry_does_not_advertise_an_unbundled_builtin() -> None:
    with TemporaryDirectory() as temporary:
        assert ModelRegistry(Path(temporary)).list() == []


def test_health_and_capabilities_contract() -> None:
    with TestClient(app) as client:
        health = client.get("/api/health")
        assert health.status_code == 200
        assert health.headers["cache-control"] == "no-store"
        assert health.json()["status"] == "ok"
        capabilities = client.get("/api/capabilities")
        assert capabilities.status_code == 200
        payload = capabilities.json()
        assert payload["cuda_required"] is True
        assert "cuda_error" in payload
        assert "formal_lite_ready" in payload
        assert payload["supported_decision_contracts"] == ["stable_advantage_v2"]


def test_schema_v3_envelope_persists_formal_runtime_identity() -> None:
    run_id = uuid4()
    result = StatisticsService.envelope(
        run_id=run_id,
        created_at=job_manager_module.utc_now(),
        config={"engine": "lite", "decision_contract": "stable_advantage_v2"},
        raw_result={
            "metrics_version": 2,
            "decision_contract": "stable_advantage_v2",
            "runtime": FORMAL_RUNTIME,
            "candidates": [],
        },
        gpu_telemetry=None,
    )
    assert result["schema_version"] == 3
    assert result["metrics_version"] == 2
    assert result["decision_contract"] == "stable_advantage_v2"
    assert result["runtime"] == FORMAL_RUNTIME


def test_run_request_accepts_public_payload() -> None:
    request = RunRequest(
        hand="4567m3477p134066s",
        dora="9s",
        discards=["1s", "6s"],
        runs=100,
        seed=42,
        round="S3",
        honba=2,
        kyotaku=1,
        scores={"self": 30000, "shimocha": 24000, "toimen": 22000},
        batch_size=100,
        model_id="mortal-v4-20240308",
        rayon_threads=4,
        engine="python",
        decision_contract="legacy_amp_v1",
    )
    assert request.model_dump()["strict_comparison"] is True
    assert request.first_tsumo is None
    assert request.model_id == "mortal-v4-20240308"
    context = resolve_simulation_context(request.model_dump())
    assert context == {
        "round": "S3",
        "kyoku": 7,
        "bakaze": "S",
        "oya": 2,
        "honba": 2,
        "kyotaku": 1,
        "scores": [22000, 23000, 30000, 24000],
        "relative_scores": {
            "self": 30000,
            "shimocha": 24000,
            "toimen": 22000,
            "kamicha": 23000,
        },
    }


def test_gpu_only_runtime_rejects_cpu(monkeypatch) -> None:
    monkeypatch.setattr(main_module, "require_cuda", lambda: (_ for _ in ()).throw(RuntimeError("CUDA unavailable")))
    payload = {
        "hand": "4567m3477p134066s",
        "dora": "9s",
        "discards": ["1s"],
    }
    with TestClient(app) as client:
        response = client.post("/api/runs", json=payload)
    assert response.status_code == 503
    assert "CUDA unavailable" in response.json()["detail"]


def test_new_hand_input_requires_fourteen_tiles_and_keeps_legacy_records_compatible() -> None:
    with pytest.raises(ValidationError, match="exactly 14 tiles"):
        RunRequest(hand="4567m3477p13406s", dora="9s", discards=["1s"])
    legacy = RunRequest(hand="4567m3477p13406s", first_tsumo="6s", dora="9s", discards=["1s"])
    assert legacy.first_tsumo == "6s"


def test_formal_lite_contract_fixes_public_batch_and_engine_pairing() -> None:
    base = {"hand": "4567m3477p134066s", "dora": "9s", "discards": ["1s"]}
    with pytest.raises(ValidationError, match="batch_size=1000"):
        RunRequest(**base, batch_size=999)
    with pytest.raises(ValidationError, match="legacy_amp_v1"):
        RunRequest(**base, engine="python")
    legacy = RunRequest(
        **base,
        engine="python",
        decision_contract="legacy_amp_v1",
        batch_size=32,
    )
    assert legacy.decision_contract == "legacy_amp_v1"


def test_fourteenth_tile_is_derived_as_first_tsumo() -> None:
    hand, first_tsumo, dora, discards, *_ = _parse_inputs({
        "hand": "4567m3477p134066s",
        "dora": "9s",
        "discards": ["1s", "6s"],
        "runs": 1,
        "seed": 42,
        "batch_size": 1,
        "rayon_threads": 1,
        "round": "E1",
        "honba": 0,
        "kyotaku": 0,
        "scores": {"self": 25000, "shimocha": 25000, "toimen": 25000},
    })
    assert hand == ["4m", "5m", "6m", "7m", "3p", "4p", "7p", "7p", "1s", "3s", "4s", "5sr", "6s"]
    assert first_tsumo == "6s"
    assert dora == "9s"
    assert discards == [
        {"tile": "1s", "engine_tile": "1s", "riichi": False, "candidate": "1s"},
        {"tile": "6s", "engine_tile": "6s", "riichi": False, "candidate": "6s"},
    ]


def test_compact_honor_discard_uses_rust_honor_name_without_changing_ui_identity() -> None:
    _, _, _, discards, *_ = _parse_inputs({
        "hand": "34m67p1178s123667z",
        "dora": "3z",
        "discards": ["1z", "3z"],
        "runs": 1,
        "seed": 89,
        "batch_size": 1,
        "rayon_threads": 1,
        "round": "E1",
        "honba": 0,
        "kyotaku": 0,
        "scores": {"self": 25_000, "shimocha": 25_000, "toimen": 25_000},
    })
    assert discards == [
        {"tile": "1z", "engine_tile": "E", "riichi": False, "candidate": "1z"},
        {"tile": "3z", "engine_tile": "W", "riichi": False, "candidate": "3z"},
    ]


def test_red_five_discard_keeps_public_zero_notation() -> None:
    _, _, _, discards, *_ = _parse_inputs({
        "hand": "04567m3477p13406s",
        "dora": "9s",
        "discards": ["0m", {"tile": "0m", "riichi": True}],
        "runs": 1,
        "seed": 89,
        "batch_size": 1,
        "rayon_threads": 1,
        "round": "E1",
        "honba": 0,
        "kyotaku": 0,
        "scores": {"self": 25_000, "shimocha": 25_000, "toimen": 25_000},
    })
    assert discards == [
        {"tile": "0m", "engine_tile": "5mr", "riichi": False, "candidate": "0m"},
        {"tile": "0m", "engine_tile": "5mr", "riichi": True, "candidate": "riichi:0m"},
    ]


def test_first_discard_riichi_is_a_distinct_candidate_action() -> None:
    request = RunRequest(
        hand="4567m3477p134066s",
        dora="9s",
        discards=["1s", {"tile": "1s", "riichi": True}],
    )
    assert request.model_dump()["discards"] == [
        {"tile": "1s", "riichi": False},
        {"tile": "1s", "riichi": True},
    ]

    with pytest.raises(ValidationError, match="unique by tile and riichi mode"):
        RunRequest(
            hand="4567m3477p134066s",
            dora="9s",
            discards=[{"tile": "1s", "riichi": True}, {"tile": "1s", "riichi": True}],
        )


@pytest.mark.parametrize(
    ("round_id", "kyoku", "bakaze", "oya"),
    [(f"{wind}{number}", wind_index * 4 + number, wind, number - 1)
     for wind_index, wind in enumerate("ESW") for number in range(1, 5)],
)
def test_all_rounds_resolve_to_global_kyoku_and_dealer(round_id: str, kyoku: int, bakaze: str, oya: int) -> None:
    request = RunRequest(
        hand="4567m3477p13406s",
        first_tsumo="6s",
        dora="9s",
        discards=["1s"],
        round=round_id,
    )
    context = resolve_simulation_context(request.model_dump())
    assert (context["kyoku"], context["bakaze"], context["oya"]) == (kyoku, bakaze, oya)
    assert context["scores"][oya] == 25_000


@pytest.mark.parametrize(
    "scores,kyotaku",
    [
        ({"self": 25001, "shimocha": 25000, "toimen": 25000}, 0),
        ({"self": -100, "shimocha": 25000, "toimen": 25000}, 0),
        ({"self": 50000, "shimocha": 50000, "toimen": 100}, 0),
        ({"self": 25000, "shimocha": 25000, "toimen": 25000}, 26),
    ],
)
def test_invalid_relative_scores_are_rejected(scores: dict[str, int], kyotaku: int) -> None:
    with pytest.raises(ValidationError):
        RunRequest(
            hand="4567m3477p13406s",
            first_tsumo="6s",
            dora="9s",
            discards=["1s"],
            scores=scores,
            kyotaku=kyotaku,
        )


def test_completed_run_metadata_survives_manager_restart() -> None:
    with TemporaryDirectory() as directory:
        first = JobManager(data_dir=Path(directory))
        job = Job(run_id=uuid4(), request={"discards": ["1s"]}, status="completed")
        job.result = {"schema_version": 1, "candidates": []}
        first.jobs[job.run_id] = job
        first._persist(job)

        second = JobManager(data_dir=Path(directory))
        restored = second.get(job.run_id)
        assert restored.status == "completed"
        assert restored.result == job.result


def test_active_run_is_marked_interrupted_after_restart() -> None:
    with TemporaryDirectory() as directory:
        first = JobManager(data_dir=Path(directory))
        job = Job(run_id=uuid4(), request={"discards": ["1s"]}, status="running")
        first.jobs[job.run_id] = job
        first._persist(job)
        restored = JobManager(data_dir=Path(directory)).get(job.run_id)
        assert restored.status == "interrupted"
        assert "restarted" in (restored.error or "")


def test_worker_failure_preserves_detail_and_writes_traceback() -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        job = Job(run_id=uuid4(), request={"discards": ["1s"]}, status="running")
        manager.jobs[job.run_id] = job
        trace = "Traceback (most recent call last):\n  ...\nTypeError: unexpected keyword argument 'first_riichi'\n"

        manager._handle_event(job, {
            "type": "failed",
            "error": "TypeError: unexpected keyword argument 'first_riichi'",
            "traceback": trace,
        })

        assert job.status == "failed"
        assert job.error == "TypeError: unexpected keyword argument 'first_riichi'"
        assert job.diagnostic_log == f"worker-{job.run_id}.log"
        assert trace in (manager.logs_dir / job.diagnostic_log).read_text(encoding="utf-8")
        persisted = manager.record(job)
        assert persisted["diagnostic_log"] == job.diagnostic_log


def test_result_exports(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        job = Job(run_id=uuid4(), request={"discards": ["1s"]}, status="completed")
        job.result = {
            "run_id": str(job.run_id),
            "summaries": [{"discard": "1s", "games": 1, "avg_point": 100, "avg_rank": 2.0}],
        }
        manager.jobs[job.run_id] = job
        monkeypatch.setattr(main_module, "manager", manager)
        with TestClient(main_module.app) as client:
            full_response = client.get(f"/api/runs/{job.run_id}/export?format=full")
            csv_response = client.get(f"/api/runs/{job.run_id}/export?format=csv")
            xlsx_response = client.get(f"/api/runs/{job.run_id}/export?format=xlsx")
            html_response = client.get(f"/api/runs/{job.run_id}/export?format=html")
        assert full_response.status_code == 200
        assert full_response.json()["run_id"] == str(job.run_id)
        assert "-full.json" in full_response.headers["content-disposition"]
        assert csv_response.status_code == 200
        assert "discard,games" in csv_response.text
        assert xlsx_response.status_code == 200
        assert xlsx_response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
        assert workbook.sheetnames == ["指标总表"]
        assert html_response.status_code == 200
        assert "平均局收支" in html_response.text
        assert "实际净点差" not in html_response.text


def test_sample_drilldown_contract(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        job = Job(run_id=uuid4(), request={"discards": ["1s"]}, status="completed")
        job.result = {
            "schema_version": 2,
            "candidates": [{
                "discard": "1s",
                "samples": {"outcome.sideways": [{"seed": [42, 57005], "outcome": "sideways"}]},
            }],
        }
        manager.jobs[job.run_id] = job
        monkeypatch.setattr(main_module, "manager", manager)
        with TestClient(main_module.app) as client:
            response = client.get(
                f"/api/runs/{job.run_id}/samples",
                params={"candidate": "1s", "metric": "outcome.sideways"},
            )
        assert response.status_code == 200
        assert response.json()["samples"][0]["outcome"] == "sideways"


def test_extension_inherits_configuration_and_uses_next_seed(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        monkeypatch.setattr(
            job_manager_module,
            "ModelRegistry",
            lambda _data_dir: type("ReadyRegistry", (), {"get": lambda self, _model_id: {"sha256": "test"}})(),
        )
        parent = Job(
            run_id=uuid4(),
            request={
                "discards": ["1s", "6s"],
                "seed": 42,
                "runs": 1000,
                "round": "S2",
                "batch_size": 1000,
                "decision_contract": "stable_advantage_v2",
            },
            status="completed",
            result={
                "schema_version": 3,
                "metrics_version": 3,
                "decision_contract": "stable_advantage_v2",
                "merge_state_version": 3,
                "runtime": FORMAL_RUNTIME,
                "model": {"sha256": "test"},
                "seed": 42,
                "runs": 1000,
                "total_runs": 1000,
            },
        )
        manager.jobs[parent.run_id] = parent
        captured: dict = {}

        def fake_create(request: dict, extension_of=None):
            captured.update(request)
            return Job(run_id=uuid4(), request=request, status="running", extension_of=extension_of)

        monkeypatch.setattr(manager, "_create_job", fake_create)
        operation = manager.create_extension(parent.run_id, 500, batch_size=1000)
        assert operation.extension_of == parent.run_id
        assert captured["round"] == "S2"
        assert captured["runs"] == 500
        assert captured["seed"] == 1042
        assert captured["batch_size"] == 1000


def test_run_extension_after_candidate_extension_uses_all_result_candidates(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        monkeypatch.setattr(
            job_manager_module,
            "ModelRegistry",
            lambda _data_dir: type("ReadyRegistry", (), {"get": lambda self, _model_id: {"sha256": "test"}})(),
        )
        parent = Job(
            run_id=uuid4(),
            request={
                "discards": [
                    {"tile": "2m", "riichi": False},
                    {"tile": "5p", "riichi": False},
                ],
                "seed": 42,
                "runs": 2000,
                "round": "E1",
                "batch_size": 1000,
                "decision_contract": "stable_advantage_v2",
            },
            status="completed",
            result={
                "schema_version": 3,
                "metrics_version": 3,
                "decision_contract": "stable_advantage_v2",
                "merge_state_version": 3,
                "runtime": FORMAL_RUNTIME,
                "model": {"sha256": "test"},
                "seed": 42,
                "runs": 2000,
                "total_runs": 2000,
                "candidates": [
                    {"candidate": "2m", "discard": "2m", "first_riichi": False},
                    {"candidate": "5p", "discard": "5p", "first_riichi": False},
                    {"candidate": "4s", "discard": "4s", "first_riichi": False, "riichi": {"rate": {"count": 1, "total": 2}}},
                    {"candidate": "3m", "discard": "3m", "first_riichi": False},
                    {"candidate": "7z", "discard": "7z", "first_riichi": False},
                ],
            },
        )
        manager.jobs[parent.run_id] = parent
        captured: dict = {}

        def fake_create(request: dict, extension_of=None):
            captured.update(request)
            return Job(run_id=uuid4(), request=request, status="running", extension_of=extension_of)

        monkeypatch.setattr(manager, "_create_job", fake_create)
        operation = manager.create_extension(parent.run_id, 1000, batch_size=1000)
        assert operation.extension_of == parent.run_id
        assert captured["discards"] == [
            {"tile": "2m", "riichi": False},
            {"tile": "5p", "riichi": False},
            {"tile": "4s", "riichi": False},
            {"tile": "3m", "riichi": False},
            {"tile": "7z", "riichi": False},
        ]
        assert captured["runs"] == 1000
        assert captured["seed"] == 2042
        assert captured["extension_mode"] == "runs"


def test_candidate_extension_after_candidate_extension_rejects_result_duplicate(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        monkeypatch.setattr(
            job_manager_module,
            "ModelRegistry",
            lambda _data_dir: type("ReadyRegistry", (), {"get": lambda self, _model_id: {"sha256": "test"}})(),
        )
        parent = Job(
            run_id=uuid4(),
            request={
                "discards": [
                    {"tile": "2m", "riichi": False},
                    {"tile": "5p", "riichi": False},
                ],
                "seed": 42,
                "runs": 2000,
                "batch_size": 1000,
                "decision_contract": "stable_advantage_v2",
            },
            status="completed",
            result={
                "schema_version": 3,
                "metrics_version": 3,
                "decision_contract": "stable_advantage_v2",
                "merge_state_version": 3,
                "runtime": FORMAL_RUNTIME,
                "model": {"sha256": "test"},
                "seed": 42,
                "runs": 2000,
                "total_runs": 2000,
                "candidates": [
                    {"candidate": "2m", "discard": "2m", "first_riichi": False},
                    {"candidate": "5p", "discard": "5p", "first_riichi": False},
                    {"candidate": "4s", "discard": "4s", "first_riichi": False, "riichi": {"rate": {"count": 1, "total": 2}}},
                ],
            },
        )
        manager.jobs[parent.run_id] = parent
        with pytest.raises(RuntimeError, match="已经存在"):
            manager.create_extension(
                parent.run_id,
                discards=[DiscardCandidate(tile="4s")],
            )


def test_extension_rejects_batch_change_to_preserve_amp_trace(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        monkeypatch.setattr(
            job_manager_module,
            "ModelRegistry",
            lambda _data_dir: type("ReadyRegistry", (), {"get": lambda self, _model_id: {"sha256": "test"}})(),
        )
        parent = Job(
            run_id=uuid4(),
            request={
                "discards": ["1s"],
                "seed": 42,
                "runs": 1000,
                "batch_size": 1000,
                "decision_contract": "stable_advantage_v2",
            },
            status="completed",
            result={
                "schema_version": 3,
                "metrics_version": 3,
                "decision_contract": "stable_advantage_v2",
                "merge_state_version": 3,
                "runtime": FORMAL_RUNTIME,
                "model": {"sha256": "test"},
                "seed": 42,
                "runs": 1000,
                "total_runs": 1000,
            },
        )
        manager.jobs[parent.run_id] = parent
        with pytest.raises(RuntimeError, match="Batch|batch"):
            manager.create_extension(parent.run_id, 100, batch_size=1)


def test_schema_v2_history_is_read_only_for_extensions() -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        parent = Job(
            run_id=uuid4(),
            request={"discards": ["1s"], "seed": 42, "runs": 1000},
            status="completed",
            result={"schema_version": 2, "metrics_version": 2, "runs": 1000},
        )
        manager.jobs[parent.run_id] = parent
        with pytest.raises(RuntimeError, match="read-only"):
            manager.create_extension(parent.run_id, 100)


def test_early_schema_v3_history_without_exact_merge_state_is_read_only() -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        parent = Job(
            run_id=uuid4(),
            request={"discards": ["1s"], "seed": 42, "runs": 1000},
            status="completed",
            result={
                "schema_version": 3,
                "metrics_version": 3,
                "merge_state_version": 1,
                "decision_contract": "stable_advantage_v2",
                "runs": 1000,
            },
        )
        manager.jobs[parent.run_id] = parent
        with pytest.raises(RuntimeError, match="精确扩容状态"):
            manager.create_extension(parent.run_id, 100)


def test_active_tasks_expose_extension_progress(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        job = Job(
            run_id=uuid4(),
            request={"discards": ["1s"], "batch_size": 12},
            status="running",
            extension_of=uuid4(),
            progress={"discard": "1s", "completed": 120, "total": 500},
            gpu_status={"type": "gpu_status", "sample": {"temperature.gpu": 66}},
        )
        manager.jobs[job.run_id] = job
        monkeypatch.setattr(main_module, "manager", manager)
        with TestClient(main_module.app) as client:
            response = client.get("/api/tasks/active")
        assert response.status_code == 200
        payload = response.json()[0]
        assert payload["extension_of"] == str(job.extension_of)
        assert payload["progress"]["completed"] == 120


def test_failed_extension_merge_leaves_parent_unchanged(monkeypatch) -> None:
    with TemporaryDirectory() as directory:
        manager = JobManager(data_dir=Path(directory))
        original = {"schema_version": 2, "metrics_version": 2, "runs": 100, "candidates": []}
        parent = Job(run_id=uuid4(), request={"discards": ["1s"]}, status="completed", result=original.copy())
        child = Job(run_id=uuid4(), request={"discards": ["1s"]}, status="running", extension_of=parent.run_id)
        manager.jobs[parent.run_id] = parent
        manager.jobs[child.run_id] = child
        monkeypatch.setattr(job_manager_module, "merge_results", lambda *_args: (_ for _ in ()).throw(ValueError("bad merge")))
        manager._handle_event(child, {"type": "completed", "result": {"candidates": []}})
        assert child.status == "failed"
        assert parent.result == original


def test_complete_public_domain_tile_set_is_packaged() -> None:
    tile_dir = Path(__file__).parents[2] / "apps" / "web" / "public" / "tiles"
    expected = {
        *(f"Man{number}.webp" for number in range(1, 10)),
        *(f"Pin{number}.webp" for number in range(1, 10)),
        *(f"Sou{number}.webp" for number in range(1, 10)),
        "Man5-Dora.webp", "Pin5-Dora.webp", "Sou5-Dora.webp",
        "Ton.webp", "Nan.webp", "Shaa.webp", "Pei.webp",
        "Haku.webp", "Hatsu.webp", "Chun.webp",
    }
    assert {path.name for path in tile_dir.glob("*.webp")} == expected
    assert "CC0" in (tile_dir / "LICENSE-FluffyStuff.md").read_text(encoding="utf-8")
